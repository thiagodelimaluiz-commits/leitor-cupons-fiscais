import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from PIL import Image
import io
import json
import google.generativeai as genai

# Configuração da página (Layout otimizado para celular)
st.set_page_config(
    page_title="Leitor de Cupons Fiscais",
    page_icon="🧾",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# Estilização CSS
st.markdown("""
<style>
    .main-header {
        font-size: 1.8rem;
        color: #1F4E79;
        font-weight: 700;
        margin-bottom: 0.3rem;
        text-align: center;
    }
    .sub-header {
        font-size: 0.95rem;
        color: #555555;
        margin-bottom: 1.5rem;
        text-align: center;
    }
    .stButton>button {
        background-color: #1F4E79;
        color: white;
        border-radius: 8px;
        font-weight: 600;
        width: 100%;
        padding: 0.8rem;
        font-size: 1.1rem;
    }
    .stButton>button:hover {
        background-color: #2F5597;
        color: white;
    }
</style>
""", unsafe_allow_html=True)

st.markdown('<div class="main-header">🧾 Leitor de Cupons Fiscais</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-header">Tire fotos pelo celular e obtenha sua planilha Excel pronta instantaneamente.</div>', unsafe_allow_html=True)

# Busca segura de Secrets sem quebrar o app
default_api_key = ""
try:
    if "GEMINI_API_KEY" in st.secrets:
        default_api_key = st.secrets["GEMINI_API_KEY"]
except Exception:
    pass

# Chave de API na barra lateral ou via Secrets do Streamlit
with st.sidebar:
    st.header("⚙️ Configurações")
    api_key = st.text_input("Chave de API do Google Gemini", type="password", value=default_api_key)
    st.info("💡 Dica: Configure sua GEMINI_API_KEY nos Secrets do Streamlit para salvar a chave permanentemente.")

# Upload com suporte à câmera do celular
uploaded_files = st.file_uploader(
    "📸 Tire foto ou selecione imagens dos cupons:", 
    type=["jpg", "jpeg", "png", "webp"], 
    accept_multiple_files=True
)

def extract_data_with_gemini(image_bytes, filename, key):
    try:
        genai.configure(api_key=key)
        
        # Tentativa de modelos na ordem de preferência/disponibilidade
        model_names = ['gemini-2.5-flash', 'gemini-1.5-flash-latest', 'gemini-2.0-flash']
        model = None
        
        for name in model_names:
            try:
                model = genai.GenerativeModel(name)
                break
            except Exception:
                continue
                
        if not model:
            model = genai.GenerativeModel('gemini-2.5-flash')

        image = Image.open(io.BytesIO(image_bytes))
        
        prompt = """
        Analise a imagem deste cupom fiscal ou recibo e extraia as seguintes informações no formato JSON exato:
        {
          "descricao_uso": "Classificação/Categoria breve do gasto (ex: Alimentação, Hardware, Material de Escritório, Combustível, etc)",
          "fornecedor": "Nome do Estabelecimento / Razão Social",
          "cidade": "Cidade e UF do estabelecimento (ex: Curitiba/PR)",
          "data_emissao": "Data de emissão formato DD/MM/AAAA",
          "horario_emissao": "Horário formato HH:MM:SS (ou HH:MM)",
          "valor_total": float do valor total (ex: 320.00),
          "produtos": "Lista legível dos produtos com quantidade e valor unitário"
        }
        Responda APENAS o JSON, sem marcações markdown de código.
        """
        
        response = model.generate_content([prompt, image])
        clean_text = response.text.replace("```json", "").replace("```", "").strip()
        data = json.loads(clean_text)
        
        return {
            "Nome do Arquivo": filename,
            "Descrição do Uso": data.get("descricao_uso", "Despesa Diversa"),
            "Fornecedor / Estabelecimento": data.get("fornecedor", "Não identificado"),
            "Cidade": data.get("cidade", "Não informada"),
            "Data de Emissão": data.get("data_emissao", "Não informada"),
            "Horário de Emissão": data.get("horario_emissao", "Não informado"),
            "Valor Total (R$)": float(data.get("valor_total", 0.0)),
            "Produtos": data.get("produtos", "Não detalhado")
        }
    except Exception as e:
        return {
            "Nome do Arquivo": filename,
            "Descrição do Uso": "Erro na leitura",
            "Fornecedor / Estabelecimento": f"Erro: {str(e)}",
            "Cidade": "-",
            "Data de Emissão": "-",
            "Horário de Emissão": "-",
            "Valor Total (R$)": 0.0,
            "Produtos": "-"
        }

def generate_excel(df_data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cupons Fiscais"
    ws.views.sheetView[0].showGridLines = True

    # Bloco do Título
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = "RELATÓRIO DE EXTRAÇÃO DE CUPONS FISCAIS"
    title_cell.font = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    headers = [
        "Nome do Arquivo", "Descrição do Uso", "Fornecedor / Estabelecimento",
        "Cidade", "Data de Emissão", "Horário de Emissão", "Valor Total (R$)", "Produtos"
    ]

    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border_side = Side(border_style="thin", color="D9D9D9")
    thick_bottom_side = Side(border_style="medium", color="1F4E79")
    border_header = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thick_bottom_side)

    ws.row_dimensions[3].height = 28

    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header_title
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border_header

    zebra_fill = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    border_cell = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    data_font = Font(name="Segoe UI", size=10)

    start_row = 4
    for i, row in df_data.iterrows():
        current_row = start_row + i
        fill = zebra_fill if i % 2 == 1 else white_fill
        ws.row_dimensions[current_row].height = 35
        
        try:
            val_total = float(row["Valor Total (R$)"])
        except:
            val_total = 0.0

        row_values = [
            (row["Nome do Arquivo"], Alignment(horizontal="center", vertical="center")),
            (row["Descrição do Uso"], Alignment(horizontal="left", vertical="center")),
            (row["Fornecedor / Estabelecimento"], Alignment(horizontal="left", vertical="center", wrap_text=True)),
            (row["Cidade"], Alignment(horizontal="center", vertical="center")),
            (row["Data de Emissão"], Alignment(horizontal="center", vertical="center")),
            (row["Horário de Emissão"], Alignment(horizontal="center", vertical="center")),
            (val_total, Alignment(horizontal="right", vertical="center")),
            (row["Produtos"], Alignment(horizontal="left", vertical="center", wrap_text=True)),
        ]
        
        for col_num, (val, align) in enumerate(row_values, 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.value = val
            cell.font = data_font
            cell.fill = fill
            cell.alignment = align
            cell.border = border_cell
            if col_num == 7:
                cell.number_format = 'R$ #,##0.00'

    # Linha do Total
    total_row = start_row + len(df_data)
    ws.row_dimensions[total_row].height = 25
    ws.cell(row=total_row, column=6).value = "Total Geral:"
    ws.cell(row=total_row, column=6).font = Font(name="Segoe UI", size=11, bold=True)
    ws.cell(row=total_row, column=6).alignment = Alignment(horizontal="right", vertical="center")

    total_cell = ws.cell(row=total_row, column=7)
    total_cell.value = f"=SUM(G{start_row}:G{total_row-1})"
    total_cell.font = Font(name="Segoe UI", size=11, bold=True, color="1F4E79")
    total_cell.number_format = 'R$ #,##0.00'
    total_cell.alignment = Alignment(horizontal="right", vertical="center")
    total_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    total_border = Border(top=Side(border_style="thin", color="1F4E79"), bottom=Side(border_style="double", color="1F4E79"))
    total_cell.border = total_border
    ws.cell(row=total_row, column=6).border = total_border

    column_widths = {"A": 16, "B": 30, "C": 40, "D": 22, "E": 18, "F": 18, "G": 18, "H": 50}
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

if uploaded_files:
    if not api_key:
        st.warning("⚠️ Forneça uma chave de API do Gemini na barra lateral para processar as imagens.")
    else:
        if st.button("🚀 Processar e Gerar Planilha"):
            extracted_rows = []
            progress_bar = st.progress(0)
            
            for idx, file in enumerate(uploaded_files):
                img_bytes = file.read()
                data_dict = extract_data_with_gemini(img_bytes, file.name, api_key)
                extracted_rows.append(data_dict)
                progress_bar.progress((idx + 1) / len(uploaded_files))
                
            st.session_state["data_df"] = pd.DataFrame(extracted_rows)
            st.success("✅ Cupons processados com sucesso!")

if "data_df" in st.session_state:
    st.markdown("### 📝 Dados Extraídos")
    edited_df = st.data_editor(st.session_state["data_df"], num_rows="dynamic", use_container_width=True)
    
    excel_bytes = generate_excel(edited_df)
    
    st.download_button(
        label="📥 Baixar Planilha Excel (.xlsx)",
        data=excel_bytes,
        file_name="Cupons_Fiscais_Extraidos.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
